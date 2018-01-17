# -*- coding: utf-8 -*-
"""
Created on Wed Jan 17 01:27:30 2018

"""

from openpyxl import Workbook
from openpyxl import load_workbook
import calendar
import random

filepath = "C:\\Users\\RPS\\Desktop\Report.xlsx"

wb=Workbook()
i=0

for month in calendar.month_name:
    if not(len(str(month))==0):
        print (month)
        wb.create_sheet(month+"_2018",i)
        i+=1
wb.save(filepath)

fileRef = load_workbook(filepath, read_only=False)

sheet = fileRef.get_sheet_by_name("March_2018")

for row in range(51, 100):
    for col in range(1,6):
        sheet.cell(column=col,row=row, 
                   value="%d" % random.randint(1,10000))
        
fileRef.save(filepath)
