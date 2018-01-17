# -*- coding: utf-8 -*-
"""
Created on Wed Jan 17 02:55:51 2018

"""

from openpyxl import  load_workbook
from openpyxl.chart import(AreaChart,
    Reference,
    Series,
    BarChart3D)
filepath="C:\\Users\\RPS\\Desktop\Report.xlsx"
wb=load_workbook(filepath, read_only=False,data_only=True)
sheet = wb.get_sheet_by_name('March_2018')
for row in range(2,sheet.max_row):
    for col in range(3,6):
        print(sheet.cell(row=row,column=col).value)
        
        
chart=BarChart3D()
chart.title="Bandwidth Utilization"
chart.style=13
chart.x_axis.title="Time"
chart.y_axis.title="Usage"
xdata=Reference(sheet, min_col=3, min_row=1, max_row=7)
ydata=Reference(sheet, min_col=4, min_row=1, 
                max_row=7,max_col=5)
chart.add_data(ydata,titles_from_data=True)
chart.set_categories(xdata)
sheet.add_chart(chart,"G15")

wb.save(filepath)